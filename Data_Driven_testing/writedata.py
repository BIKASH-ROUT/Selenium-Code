import openpyxl

# Write file location which having blank sheet
file = "C:\\Users\\routb\\OneDrive\\Documents\\DDT(write).xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook.active            # = Blank current page

# # For same value in every row and column
# for r in range(1,5):
#     for c in range(1,3):
#         sheet.cell(r,c).value="welcome"
#
# workbook.save(file)

### 1st clear all data from sheet or create new sheet
# Multiple Data
sheet.cell(1,1).value=2
sheet.cell(1,2).value="Bikash"
sheet.cell(1,3).value="Engg"

sheet.cell(2,1).value=2
sheet.cell(2,2).value="John"
sheet.cell(2,3).value="Manager"

sheet.cell(3,1).value=3
sheet.cell(3,1).value="Smith"
sheet.cell(3,1).value="tester"

workbook.save()

