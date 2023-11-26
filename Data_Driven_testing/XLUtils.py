import openpyxl
from openpyxl.styles import PatternFill

# Row Count :
def getRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_row)

# Column Count:
def getColumnCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_column)

# Read Data
def readData(file,sheetName,rownum,columnnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(rownum,columnnum).value

# Write Data
def writeData(file,sheetName,rownum,columnnum,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(rownum,columnnum).value = data
    workbook.save(file)

# Custome green color:
def fillGreenColor(file,sheetName,rownum,columnnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    greenFill = PatternFill(start_color="60b212",
                            end_color="60b212",
                            fill_type="solid")
    sheet.cell(rownum,columnnum).fill = greenFill
    workbook.save(file)

# Custom Red color
def fillRedColor(file,sheetName,rownum,columnnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    redFill = PatternFill(start_color="ff0000",
                            end_color="ff0000",
                            fill_type="solid")
    sheet.cell(rownum,columnnum).fill = redFill
    workbook.save(file)